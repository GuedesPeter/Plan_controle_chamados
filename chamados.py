from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# =================================================
# CONFIGURAÇÕES
# =================================================
funcionarios = ["Douglas", "Gerson", "Paulo"]

headers = [
    "Número do Chamado",
    "Status",
    "Descrição",
    "Início EM",
    "Finalizar?"
]

status_list = [
    "Novo",
    "Em Atendimento",
    "Pendente",
    "Solucionado",
    "Finalizado"
]

status_colors = {
    "Novo": "2E7D32",
    "Em Atendimento": "66BB6A",
    "Pendente": "EF6C00",
    "Solucionado": "1565C0",
    "Finalizado": "424242"
}

AZUL_CLARO_LINHA = "D9EAF7"

# =================================================
# WORKBOOK
# =================================================
wb = Workbook()
wb.remove(wb.active)

# =================================================
# FORMATAÇÃO PADRÃO
# =================================================
def formatar_aba(ws):
    for col in range(1, len(headers) + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 30
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="ECECEC")

# =================================================
# ABAS DOS FUNCIONÁRIOS
# =================================================
for nome in funcionarios:
    ws = wb.create_sheet(title=nome)
    ws.append(headers)
    formatar_aba(ws)

    for row in range(2, 1001):
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    dv_status = DataValidation(
        type="list",
        formula1=f'"{",".join(status_list)}"'
    )
    ws.add_data_validation(dv_status)
    dv_status.add("B2:B1000")

    dv_finalizar = DataValidation(
        type="list",
        formula1='"SIM,NÃO"'
    )
    ws.add_data_validation(dv_finalizar)
    dv_finalizar.add("E2:E1000")

    for row in range(2, 1001):
        for status, color in status_colors.items():
            ws.conditional_formatting.add(
                f"B{row}",
                FormulaRule(
                    formula=[f'B{row}="{status}"'],
                    fill=PatternFill("solid", fgColor=color),
                    font=Font(bold=True, color="FFFFFF")
                )
            )

        ws.conditional_formatting.add(
            f"A{row}:E{row}",
            FormulaRule(
                formula=[f'$E{row}="SIM"'],
                fill=PatternFill("solid", fgColor=AZUL_CLARO_LINHA)
            )
        )

# =================================================
# ABA RESUMO / DASHBOARD
# =================================================
resumo = wb.create_sheet("Resumo")

# ---------- TABELA BASE (GERAL) ----------
resumo.append(["Status", "Geral", "Douglas", "Gerson", "Paulo"])

for col in range(1, 6):
    cell = resumo.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill("solid", fgColor="ECECEC")

for i, status in enumerate(status_list, start=2):
    resumo[f"A{i}"] = status
    resumo[f"A{i}"].fill = PatternFill("solid", fgColor=status_colors[status])
    resumo[f"A{i}"].font = Font(bold=True, color="FFFFFF")
    resumo[f"A{i}"].alignment = Alignment(horizontal="center")

    resumo[f"B{i}"] = (
        f'=COUNTIF(Douglas!B:B,A{i})'
        f'+COUNTIF(Gerson!B:B,A{i})'
        f'+COUNTIF(Paulo!B:B,A{i})'
    )
    resumo[f"C{i}"] = f'=COUNTIF(Douglas!B:B,A{i})'
    resumo[f"D{i}"] = f'=COUNTIF(Gerson!B:B,A{i})'
    resumo[f"E{i}"] = f'=COUNTIF(Paulo!B:B,A{i})'

    for col in ["B", "C", "D", "E"]:
        resumo[f"{col}{i}"].alignment = Alignment(horizontal="center")

# =================================================
# FUNÇÃO PARA CRIAR PIZZA
# =================================================
def criar_pizza(coluna, titulo, posicao):
    chart = PieChart()
    chart.title = titulo

    labels = Reference(resumo, min_col=1, min_row=2, max_row=6)
    data = Reference(resumo, min_col=coluna, min_row=1, max_row=6)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    for i, status in enumerate(status_list):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = status_colors[status]
        chart.series[0].data_points.append(dp)

    resumo.add_chart(chart, posicao)

# =================================================
# GRÁFICOS DE PIZZA
# =================================================
criar_pizza(2, "Chamados por Status - Geral", "G2")
criar_pizza(3, "Chamados por Status - Douglas", "G20")
criar_pizza(4, "Chamados por Status - Gerson", "M20")
criar_pizza(5, "Chamados por Status - Paulo", "S20")

# =================================================
# SALVAR
# =================================================
wb.save("controle_chamados.xlsx")
print("✔ Planilha 'controle_chamados.xlsx' criada com sucesso!")
